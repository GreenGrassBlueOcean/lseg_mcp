#!/usr/bin/env python3
r"""
Python-side harvester for LSEG Workspace exports (Data Item Browser + Screener).

Parses Excel files exported from:
  - Data Item Browser (DIB) — tables or lists of fields
  - Screener → "Export All as Formulas" (the goldmine — emits the exact TR.* codes)

It produces a clean CSV in the exact flat format expected by lseg-mcp's
DataDictionary (the one used by `search_data_dictionary` and `draft_api_call`).

You can then either:
  - Point LSEG_DATA_DICTIONARY_PATH at the CSV, or
  - Copy the rows into a sheet named "Custom_Fields" / "Data Dictionary" inside LSEG_Mapping.xlsx

Usage (from the repo root):

    # Basic auto-scan of a Screener or DIB export
    python scripts/harvest_lseg_fields.py -i my_screener_export.xlsx -c Estimates -o my_estimates.csv

    # Only scan a specific sheet
    python scripts/harvest_lseg_fields.py -i dib_export.xlsx -s "My Fields" -c Pricing

    # Use as a library from a notebook / another script
    from scripts.harvest_lseg_fields import parse_excel_for_tr_fields, build_field_catalog, write_for_lseg_mcp
    fields = parse_excel_for_tr_fields("export.xlsx")
    cat = build_field_catalog(fields, category="ESG")
    write_for_lseg_mcp(cat, "esg_from_dib.csv")

Dependencies: pandas + openpyxl (already required by lseg-mcp).
"""

from __future__ import annotations

import argparse
import re
from pathlib import Path
from typing import Iterable

import pandas as pd
from openpyxl import load_workbook

# Robust patterns for TR. fields coming from Workspace exports.
TR_TOKEN_RE = re.compile(r"\b(TR\.[A-Za-z0-9_]+(?:\([^)]*\))?)\b", re.IGNORECASE)


def extract_tr_fields(text: str | None) -> list[str]:
    """
    Extract TR.* fields (keeping parameters when present) from any text cell
    or formula.

    This is deliberately forgiving so it works on:
      - Screener "Export All as Formulas" (fields inside RDP.Data quoted arg)
      - DIB tables / copied lists
      - Any cell that happens to mention a TR. item
    """
    if not text or not isinstance(text, str):
        return []

    found: list[str] = []

    # Strategy: find every quoted string that contains "TR.", then split it.
    # This catches the fields argument in =RDP.Data("ric","TR.Foo,TR.Bar(...)") style exports.
    for q in re.findall(r'["\']([^"\']*TR\.[^"\']*)["\']', text, flags=re.IGNORECASE):
        for part in re.split(r"\s*,\s*", q):
            part = part.strip()
            if part.upper().startswith("TR."):
                found.append(part)

    # Also catch any bare TR. tokens that are not inside quotes (plain DIB lists, etc.)
    for m in TR_TOKEN_RE.findall(text):
        m = m.strip()
        if m.upper().startswith("TR.") and m not in found:
            found.append(m)

    # Deduplicate preserving order
    seen: set[str] = set()
    out: list[str] = []
    for f in found:
        if f not in seen:
            seen.add(f)
            out.append(f)
    return out


def parse_excel_for_tr_fields(
    path: str | Path,
    sheet_name: str | None = None,
    mode: str = "auto",
) -> list[str]:
    """
    Scan an Excel workbook for TR. fields.

    Parameters
    ----------
    path
        Path to .xlsx exported from DIB or Screener.
    sheet_name
        Optional specific sheet. If None, scans all sheets.
    mode
        "auto" (default), "screener", or "dib". Currently mostly affects internal heuristics.

    Returns
    -------
    list[str]
        Unique TR. fields found (in order of first appearance).
    """
    path = Path(path).expanduser().resolve()
    if not path.exists():
        raise FileNotFoundError(f"Input file not found: {path}")

    all_fields: list[str] = []

    # Use openpyxl with data_only=False so we see the *formula text* (critical
    # for Screener "Export All as Formulas" exports). This is more reliable
    # than pandas for our use case.
    wb = load_workbook(path, data_only=False)
    sheets = [wb[sheet_name]] if sheet_name and sheet_name in wb.sheetnames else wb.worksheets
    for ws in sheets:
        for row in ws.iter_rows(values_only=False):
            for cell in row:
                val = cell.value
                if val is not None:
                    all_fields.extend(extract_tr_fields(str(val)))

    # Deduplicate preserving first-seen order
    seen = set()
    unique: list[str] = []
    for f in all_fields:
        if f not in seen:
            seen.add(f)
            unique.append(f)
    return unique


def extract_parameters(field: str) -> str:
    """If the field contains (Params), return just the inside; otherwise ''."""
    m = re.search(r"\(([^)]+)\)", field)
    return m.group(1).strip() if m else ""


def strip_parameters(field: str) -> str:
    """Return the base TR.XXX without the (params) suffix."""
    return re.sub(r"\(.*\)$", "", field).strip()


def build_field_catalog(
    fields: Iterable[str],
    category: str = "General",
    descriptions: list[str] | None = None,
    notes: str = "Harvested via Python harvester from Workspace DIB/Screener export",
) -> pd.DataFrame:
    """
    Build the canonical 5-column DataFrame that lseg-mcp's DataDictionary understands.

    Columns: field, description, category, parameters, notes
    Duplicates on the base field name are removed (first seen wins).
    """
    fields = list(fields)
    if descriptions is None or len(descriptions) != len(fields):
        descriptions = [""] * len(fields)

    rows = []
    seen_bases: set[str] = set()
    for f, desc in zip(fields, descriptions):
        base = strip_parameters(f)
        if base in seen_bases:
            continue
        seen_bases.add(base)
        params = extract_parameters(f)
        rows.append(
            {
                "field": base,
                "description": desc,
                "category": category,
                "parameters": params,
                "notes": notes,
            }
        )
    return pd.DataFrame(rows)


def write_for_lseg_mcp(
    catalog: pd.DataFrame,
    output_path: str | Path = "harvested_lseg_fields.csv",
) -> Path:
    """
    Write the catalog to CSV using the exact columns expected by the MCP.

    The resulting file can be used directly via:
        $env:LSEG_DATA_DICTIONARY_PATH = "C:\path\to\harvested_lseg_fields.csv"
    or copied into a sheet in your LSEG_Mapping.xlsx.
    """
    cols = ["field", "description", "category", "parameters", "notes"]
    for c in cols:
        if c not in catalog.columns:
            catalog[c] = ""
    out_df = catalog[cols].copy()
    out_path = Path(output_path).expanduser().resolve()
    out_df.to_csv(out_path, index=False)
    print(f"✓ Wrote {len(out_df)} rows to {out_path}")
    return out_path


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Harvest TR.* fields from LSEG Workspace DIB or Screener Excel exports.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Screener export (most powerful)
  python scripts/harvest_lseg_fields.py -i screener_formulas.xlsx -c "Estimates" -o estimates.csv

  # DIB export on a specific sheet, assign to Pricing
  python scripts/harvest_lseg_fields.py -i dib_export.xlsx -s "Sheet1" -c "Pricing"

  # Let the script guess and just dump everything to one file
  python scripts/harvest_lseg_fields.py -i my_export.xlsx --output all_new_fields.csv
        """,
    )
    parser.add_argument(
        "-i", "--input", required=True, help="Path to the .xlsx file exported from Workspace (DIB or Screener)"
    )
    parser.add_argument(
        "-s", "--sheet", default=None, help="Specific sheet name to scan (optional; scans all if omitted)"
    )
    parser.add_argument(
        "-c", "--category", default="General", help="Category label to assign (e.g. Pricing, Estimates, ESG, Ownership)"
    )
    parser.add_argument(
        "-o", "--output", default="harvested_lseg_fields.csv", help="Output CSV path for the Data Dictionary"
    )
    parser.add_argument(
        "--mode",
        choices=["auto", "screener", "dib"],
        default="auto",
        help="Hint for the parser (currently mostly for future use / logging)",
    )
    args = parser.parse_args()

    print(f"Scanning: {args.input}")
    if args.sheet:
        print(f"Sheet filter: {args.sheet}")
    print(f"Assigning category: {args.category}")

    fields = parse_excel_for_tr_fields(args.input, sheet_name=args.sheet, mode=args.mode)

    if not fields:
        print("⚠ No TR.* fields were found. Check that you exported using 'Export All as Formulas' (Screener) or copied fields from DIB.")
        print("  Tip: In Screener, look for the 'Export' option that says 'as Formulas'.")
        return

    print(f"Found {len(fields)} unique TR. field(s).")
    for f in fields[:8]:
        print(f"  - {f}")
    if len(fields) > 8:
        print(f"  ... and {len(fields)-8} more")

    catalog = build_field_catalog(fields, category=args.category)
    write_for_lseg_mcp(catalog, args.output)

    print("\nNext steps:")
    print(f"  1. Open {args.output} and fill in better descriptions / tweak categories if needed.")
    print(f'  2. In PowerShell:  $env:LSEG_DATA_DICTIONARY_PATH = "{Path(args.output).resolve()}"')
    print("  3. Restart your MCP client (Claude Desktop, Antigravity, etc.) or call the rescan_packages tool.")
    print("  4. Ask your LLM: \"search the data dictionary for ...\" or just request code for the new fields.")


if __name__ == "__main__":
    main()
